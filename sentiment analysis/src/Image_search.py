import os, time, urllib.request, openpyxl, operator, tweepy
from tkinter import *
import tkinter as t
from openpyxl import Workbook


class TwitterScrapper:

    def gui(self, root):

        #self.root = t.Tk()
        #self.root.title("Image Search")
        #self.root.resizable(False, False)
        #self.root.config(bg = "grey")

        self.f = Frame(root, height=500, width=800)
        self.f.config(bg = "#99d8d0")
        self.f.propagate(0)
        self.f.pack()

        self.l = Label(self.f, text="Tag Based Image Search")
        self.l.config(font=("Courier", 38), bg = "#99d8d0")
        self.l.place(x=60, y=50, anchor="nw")

        self.l1 = Label(self.f, text="Enter Image Keyword :")
        self.l1.config(font=("Courier", 16), bg = "#99d8d0")
        self.l1.place(x=90, y=150, anchor='nw')

        self.e = Entry(self.f, text="Keyword", width=35, font=("Times new roman", 15))
        self.e.place(x=390, y=150)

        self.l2 = Label(self.f, text="Enter Limit :")
        self.l2.config(font=("Courier", 16), bg = "#99d8d0")
        self.l2.place(x=193, y=200, anchor="nw")

        self.e1 = Entry(self.f, text="Count", width=35, font=("Times new roman", 15))
        self.e1.place(x=390, y=200)

        #x = e.get()
        #y = e1.get()
        #print(x,y)

        self.b1 = Button(self.f, text="Search", height=2, width=10,
                     command=lambda: self.onclick())  # command =self.onclick
        self.b1.place(x=390, y=250, anchor="nw")

        self.b2 = Button(self.f, text="Exit", height=2, width=10, command=lambda: root.destroy())
        self.b2.place(x=490, y=250, anchor="nw")

    def onclick(self):
        consumerKey = 'e9RuYkFI4EEcdxWmcacjxzgkA'
        consumerSecret = 'RQWE0QSUlUXhEGxy2mVoA83DQwxqq9v5KzeqTt95vcNeLN16gn'
        accessToken = '1270244974366543874-BngXbjyLLyWkWQKW5xvPemVnL9tgzz'
        accessTokenSecret = 'cI967H4YmuMRNX8NoQZeuKbzoyhZ4lWL742HgOjvzVPEk'
        x = self.e.get()
        y = int(self.e1.get())
        #print(x,y)
        self.Scrape_Twitter(tag=x,
                           limit=y,
                           lang='en')


    def Create_Dir(self, dir_name):
        if not os.path.exists("data"):
            try:
                os.mkdir("data")
                print("Created directory 'data'")
            except:
                print("Unable to create directory 'data': Directory already exists")
        else:
            print("Unable to create directory 'data': Directory already exists")

        if not os.path.exists("data/data_" + dir_name):
            try:
                os.mkdir("data/data_" + dir_name)
                print("Created directory 'data/data_" + dir_name + "'")
            except:
                print("Unable to create directory 'data/data_" + dir_name + "': Directory already exists")
        else:
            print("Unable to create directory 'data/data_" + dir_name + "': Directory already exists")

        if not os.path.exists("data/data_" + dir_name + '/img'):
            try:
                os.mkdir("data/data_" + dir_name + '/img')
                print("Created directory 'data/data_" + dir_name + "/img'")
            except:
                print("Unable to create directory 'data/data_" + dir_name + "/img': Directory already exists")
        else:
            print("Unable to create directory 'data/data_" + dir_name + "/img': Directory already exists")

    def Scrape_Twitter(self, tag, limit=20, lang='en'):
        self.Create_Dir(tag)

        print("Starting Scrapping Twitter")
        file_path = "data/data_" + tag
        keyword = tag

        consumerKey = 'e9RuYkFI4EEcdxWmcacjxzgkA'
        consumerSecret = 'RQWE0QSUlUXhEGxy2mVoA83DQwxqq9v5KzeqTt95vcNeLN16gn'
        accessToken = '1270244974366543874-BngXbjyLLyWkWQKW5xvPemVnL9tgzz'
        accessTokenSecret = 'cI967H4YmuMRNX8NoQZeuKbzoyhZ4lWL742HgOjvzVPEk'

        auth = tweepy.OAuthHandler(consumerKey, consumerSecret)
        auth.set_access_token(accessToken, accessTokenSecret)
        api = tweepy.API(auth)

        tweets = tweepy.Cursor(api.search, q=tag, lang=lang).items(limit)

        # Create a workbook for excel
        tag_File = file_path + "/" + tag + "_Twitter.xlsx"
        wb = openpyxl.Workbook()
        ws_Captions = wb.create_sheet(title="Posts")
        col = 'A'
        row = 1

        img_src = []
        hashtags = {}
        #ext_links = []
        for tweet in tweets:
            text = tweet.text.lower()
            ws_Captions[col + str(row)] = text

            # stripping for urls and hashtags and their frequencies
            for tag in text.split():
                if tag.startswith("#"):
                    if tag[1:] not in hashtags:
                        hashtags[tag[1:]] = 1
                    elif tag[1:] in hashtags:
                        hashtags[tag[1:]] = hashtags[tag[1:]] + 1
                '''if tag[:4] == 'http':
                    ext_links.append(tag)'''

            try:
                mu = tweet.entities['media'][0]['media_url']
                if (len(mu) > 1):
                    img_src.append(mu)
            except:
                pass

            if (row % 50 == 0):
                print("Dumped " + str(row) + " Tweets")
            row += 1


        hashtags = sorted(hashtags.items(), key=operator.itemgetter(1), reverse=True)

        ws_Tags = wb.create_sheet(title="Tags")
        tagName = 'A'
        tagFreq = 'B'
        row = 1

        print("Dumping Related Hashtags")
        for tag in hashtags:
            ws_Tags[tagName + str(row)] = tag[0]
            ws_Tags[tagFreq + str(row)] = tag[1]
            row += 1


        '''print("Dumping External Links")
        ws_Links = wb.create_sheet(title="Links")
        row = 1
        for link in ext_links:
            ws_Links['A' + str(row)] = link
            row += 1'''


        wb.save(tag_File)

        time.sleep(5)

        print("Dumping " + str(len(img_src)) + " Images")
        row = 1
        for src in img_src:
            try:
                print("(" + str(row) + "/" + str(len(img_src)) + ") Images Downloaded")
                urllib.request.urlretrieve(src, file_path + '/img/Twitter_' + str(row) + ".jpeg")
                row += 1
                time.sleep(1.5)
            except:
                print("Image Download Failed. Downloading next image")

        print("Closing Twitter")



def main():
    print("Starting Social Media Scrapper...")
    #keyword = str(input("Enter keyword to search for: "))
    #insta_limit = int(input("Enter how many posts to scrape from Instagram: "))
    #twitter_limit = int(input("Enter how many posts to scrape from Twitter: "))

    '''consumerKey = 'e9RuYkFI4EEcdxWmcacjxzgkA'
    consumerSecret = 'RQWE0QSUlUXhEGxy2mVoA83DQwxqq9v5KzeqTt95vcNeLN16gn'
    accessToken = '1270244974366543874-BngXbjyLLyWkWQKW5xvPemVnL9tgzz'
    accessTokenSecret = 'cI967H4YmuMRNX8NoQZeuKbzoyhZ4lWL742HgOjvzVPEk'''

    '''twitter = TwitterScrapper()
    twitter.Scrape_Twitter(Consumer_Key=consumerKey,
                           Consumer_Secret=consumerSecret,
                           Access_Token=accessToken,
                           Access_Token_Secret=accessTokenSecret,
                           tag=keyword,
                           limit=twitter_limit,
                           lang='en')  # Language codes: https://en.wikipedia.org/wiki/List_of_ISO_639-1_codes'''


    ts = TwitterScrapper()
    root = t.Tk()
    ts.gui(root)
    root.mainloop()
    #print("Stopping Social Media Scrapper...")

#if __name__ == '__main__':
#main()